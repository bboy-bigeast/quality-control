import logging
from django.forms import model_to_dict
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse  
from openpyxl import Workbook
from django.http import JsonResponse
from django.contrib import messages
from django.db.models import Q
from .forms import DryFilmForm, AdhesiveForm, RawMaterialForm, TrialProductForm
from .models import DryFilmProduct, AdhesiveProduct, RawMaterial, TrialProduct, ReportCenter, AnalysisCenter, DryFilmStandard
from django.views.decorators.http import require_POST
from .forms import DryFilmConfigForm, DryFilmMainForm
import json
from django.core.serializers.json import DjangoJSONEncoder
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger



def home(request):
    return render(request, 'qc_app/home.html')

def base(request):
    return render(request, 'base.html')

def my_view(request):
    object_list = DryFilmProduct.objects.all().order_by('-test_date')
    paginator = Paginator(object_list, 10)  # 每页10条数据
    
    page = request.GET.get('page')
    try:
        objects = paginator.page(page)
    except PageNotAnInteger:
        objects = paginator.page(1)  # 如果page不是整数，返回第一页
    except EmptyPage:
        objects = paginator.page(paginator.num_pages)  # 超出范围返回最后一页
    
    return render(request, 'your_template.html', {'objects': objects})

#####################################################################################################
logger = logging.getLogger(__name__)

def dryfilm_list(request):
    # 获取所有干膜产品，按测试日期倒序排列
    products = DryFilmProduct.objects.all()
    # 记录查询结果
    logger.info(f"查询到 {len(products)} 条干膜产品记录")

    product_list = DryFilmProduct.objects.all().order_by('-test_date')
    paginator = Paginator(product_list, 10)  # 每页10条数据
    
    page = request.GET.get('page')
    try:
        objects = paginator.page(page)
    except PageNotAnInteger:
        objects = paginator.page(1)  # 如果page不是整数，返回第一页
    except EmptyPage:
        objects = paginator.page(paginator.num_pages)  # 超出范围返回最后一页
    
    return render(request, 'qc_app/dryfilm_list.html', {'products': products, 'objects': objects})

def dryfilm_create(request):
    if request.method == 'POST': 
        print(request.POST)
        print(request.POST['action'])
        if request.POST['action'] == 'judge':
            form = DryFilmForm(request.POST)

            if form.is_valid():
                # 步骤1：先保存用户输入
                instance = form.save()
                # 步骤2：获取刚保存的数据中的索引值,产品牌号
                target_index = instance.product_id

                try:
                    # 步骤3：从其他表提取相关数据
                    standard = DryFilmStandard.objects.get(product_name=target_index).config_json
                    print(standard)
                    # 步骤4：执行计算逻辑（示例：简单乘法计算）
                    instance_dict = model_to_dict(
                        instance,
                        fields=[field.name for field in instance._meta.fields],
                        exclude=[
                            'id', 
                            'created_at', 
                            'updated_at',
                            'batch_number', 
                            'product_id', 
                            'production_line',
                            'appearance', 
                            'inspector', 
                            'test_date', 
                            'sample_type', 
                            'remarks',
                            'factory_judgment', 
                            'internal_judgment',
                            'inhibitor', 
                            'conversion_rate', 
                            'loading_temp',
                            'change_log'
                            ]  # 排除不需要的字段
                    )

                    a = 0; b = 0; c = 0; x = 0; y = 0; z = 0

                    for i in instance_dict:
                        V = instance_dict[i]
                        print(i)
                        print(V)
                    ##################################################################
                        # 判段有无外控标准
                        if (standard[i+'_max'] != None)&(standard[i+'_min'] != None):
                            print(standard[i+'_max'])
                            print(standard[i+'_min'])
                            if isinstance(V, (int, float)) and not isinstance(V, bool):
                                if (V < float(standard[i+'_max'])) & (V > float(standard[i+'_min'])):
                                    print('OK')
                                    a+=1
                                else:
                                    print('NG')
                                    b+=1
                            else:
                                print('有未完成项')
                                c+=1
                                
                        #判断是否是大于标准值的范围
                        elif (standard[i+'_max'] == None)&(standard[i+'_min'] != None):
                            print(standard[i+'_min'])
                            if isinstance(V, (int, float)) and not isinstance(V, bool):
                                if V > float(standard[i+'_min']):
                                    print('OK')
                                    a+=1
                                else:
                                    print('NG')
                                    b+=1
                            else:
                                print('有未完成项')
                                c+=1
                        else:    
                            print('无外控标准')
                    ##################################################################
                        # 判段有无内控标准
                        if (standard['in_'+i+'_max'] != None)&(standard['in_'+i+'_min'] != None):
                            print(standard['in_'+i+'_max'])
                            print(standard['in_'+i+'_min'])
                            if isinstance(V, (int, float)) and not isinstance(V, bool):
                                if (V < float(standard['in_'+i+'_max'])) & (V > float(standard['in_'+i+'_min'])):
                                    print('OK')
                                    x+=1
                                else:
                                    print('NG')
                                    y+=1
                            else:
                                print('有未完成项')
                                z+=1
                                
                        #判断是否是大于标准值的范围
                        elif (standard['in_'+i+'_max'] == None)&(standard['in_'+i+'_min'] != None):
                            print(standard['in_'+i+'_min'])
                            if isinstance(V, (int, float)) and not isinstance(V, bool):
                                if V > float(standard['in_'+i+'_min']):
                                    print('OK')
                                    x+=1
                                else:
                                    print('NG')
                                    y+=1
                            else:
                                print('有未完成项')
                                z+=1
                        else:    
                            print('无外控标准')
                    ##################################################################
                    # 步骤5：更新原数据记录的计算结果字段
                    print('外控标准判定')
                    if (b==0):
                        if (c==0):
                            if (a>0):
                                print('判定')
                                instance.factory_judgment = 'qualified'
                        else:
                            instance.factory_judgment = 'pending'
                    else:
                        instance.factory_judgment = 'unqualified'
                    ##################################################################
                    print('内控标准判定')
                    if (y==0):
                        if (z==0):
                            if (x>0):
                                print('判定')
                                instance.internal_judgment = 'qualified'
                        else:
                            instance.internal_judgment = 'pending'
                    else:
                        instance.internal_judgment = 'unqualified'
                    
                    ##################################################################
                    instance.save(update_fields=['factory_judgment'])
                    instance.save(update_fields=['internal_judgment'])
                    print('保存')
                    ##################################################################
                    return render(request, 'qc_app/dryfilm_form.html', {'form': form})
                
                except Exception as e:
                    instance.factory_judgment = 'no_standard'
                    instance.save(update_fields=['factory_judgment'])
                    return JsonResponse({
                        'success': False,
                        'message': f'数据已经保存，未建立标准或有其他错误：{str(e)}'
                    }, status=500)

        elif request.POST['action'] == 'save':
            return redirect('dryfilm_list')
    
    else:
        form = DryFilmForm()
    
    context = {
        'form': form,
        'mode': 'create',
        'title': '添加干膜产品'
    }
    return render(request, 'qc_app/dryfilm_form.html', context)

def dryfilm_edit(request, pk):
    product = get_object_or_404(DryFilmProduct, pk=pk)
    if request.method == 'POST':
        if request.POST['action'] == 'judge':
            form = DryFilmForm(request.POST, instance=product)

            if form.is_valid():
                # 步骤1：先保存用户输入
                instance = form.save()
                # 步骤2：获取刚保存的数据中的索引值,产品牌号
                target_index = instance.product_id

                try:
                    # 步骤3：从其他表提取相关数据
                    standard = DryFilmStandard.objects.get(product_name=target_index).config_json
                    print(standard)
                    # 步骤4：执行计算逻辑（示例：简单乘法计算）
                    instance_dict = model_to_dict(
                        instance,
                        fields=[field.name for field in instance._meta.fields],
                        exclude=[
                            'id', 
                            'created_at', 
                            'updated_at',
                            'batch_number', 
                            'product_id', 
                            'production_line',
                            'appearance', 
                            'inspector', 
                            'test_date', 
                            'sample_type', 
                            'remarks',
                            'factory_judgment', 
                            'internal_judgment',
                            'inhibitor', 
                            'conversion_rate', 
                            'loading_temp',
                            'change_log'
                            ]  # 排除不需要的字段
                    )

                    a = 0; b = 0; c = 0; x = 0; y = 0; z = 0

                    for i in instance_dict:
                        V = instance_dict[i]
                        print(i)
                        print(V)
                    ##################################################################
                        # 判段有无外控标准
                        if (standard[i+'_max'] != None)&(standard[i+'_min'] != None):
                            print(standard[i+'_max'])
                            print(standard[i+'_min'])
                            if isinstance(V, (int, float)) and not isinstance(V, bool):
                                if (V < float(standard[i+'_max'])) & (V > float(standard[i+'_min'])):
                                    print('OK')
                                    a+=1
                                else:
                                    print('NG')
                                    b+=1
                            else:
                                print('有未完成项')
                                c+=1
                                
                        #判断是否是大于标准值的范围
                        elif (standard[i+'_max'] == None)&(standard[i+'_min'] != None):
                            print(standard[i+'_min'])
                            if isinstance(V, (int, float)) and not isinstance(V, bool):
                                if V > float(standard[i+'_min']):
                                    print('OK')
                                    a+=1
                                else:
                                    print('NG')
                                    b+=1
                            else:
                                print('有未完成项')
                                c+=1
                        else:    
                            print('无外控标准')
                    ##################################################################
                        # 判段有无内控标准
                        if (standard['in_'+i+'_max'] != None)&(standard['in_'+i+'_min'] != None):
                            print(standard['in_'+i+'_max'])
                            print(standard['in_'+i+'_min'])
                            if isinstance(V, (int, float)) and not isinstance(V, bool):
                                if (V < float(standard['in_'+i+'_max'])) & (V > float(standard['in_'+i+'_min'])):
                                    print('OK')
                                    x+=1
                                else:
                                    print('NG')
                                    y+=1
                            else:
                                print('有未完成项')
                                z+=1
                                
                        #判断是否是大于标准值的范围
                        elif (standard['in_'+i+'_max'] == None)&(standard['in_'+i+'_min'] != None):
                            print(standard['in_'+i+'_min'])
                            if isinstance(V, (int, float)) and not isinstance(V, bool):
                                if V > float(standard['in_'+i+'_min']):
                                    print('OK')
                                    x+=1
                                else:
                                    print('NG')
                                    y+=1
                            else:
                                print('有未完成项')
                                z+=1
                        else:    
                            print('无外控标准')
                    ##################################################################
                    # 步骤5：更新原数据记录的计算结果字段
                    print('外控标准判定')
                    if (b==0):
                        if (c==0):
                            if (a>0):
                                print('判定')
                                instance.factory_judgment = 'qualified'
                        else:
                            instance.factory_judgment = 'pending'
                    else:
                        instance.factory_judgment = 'unqualified'
                    ##################################################################
                    print('内控标准判定')
                    if (y==0):
                        if (z==0):
                            if (x>0):
                                print('判定')
                                instance.internal_judgment = 'qualified'
                        else:
                            instance.internal_judgment = 'pending'
                    else:
                        instance.internal_judgment = 'unqualified'
                    
                    ##################################################################
                    instance.save(update_fields=['factory_judgment'])
                    instance.save(update_fields=['internal_judgment'])
                    print('保存')
                    form = DryFilmForm(instance=instance)
                    ##################################################################
                    return render(request, 'qc_app/dryfilm_form.html', {'form': form,'product': instance})
                
                except Exception as e:
                    instance.factory_judgment = 'no_standard'
                    instance.save(update_fields=['factory_judgment'])
                    return JsonResponse({
                        'success': False,
                        'message': f'数据已经保存，未建立标准或有其他错误：{str(e)}'
                    }, status=500)
                
        elif request.POST['action'] == 'save':
            return redirect('dryfilm_list')
    else:
        form = DryFilmForm(instance=product)
    return render(request, 'qc_app/dryfilm_form.html', {'form': form, 'mode': 'edit', 'product': product})

@require_POST
def dryfilm_delete(request):
    record_id = request.POST.get('record_id')
    
    if not record_id:
        return JsonResponse({'success': False, 'message': '无效的请求参数'}, status=400)

    try:
        # 获取并删除记录
        record = get_object_or_404(DryFilmProduct, pk=record_id)
        record.delete()
        
        messages.success(request, f"{record.product_id}-{record.batch_number} 已成功删除")
        # 成功响应
        return redirect('dryfilm_list')
        return JsonResponse({
            'success': True, 
            'record_id': record_id,
            'message': f'批号 {record.batch_number} 已删除'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'删除失败: {str(e)}'
        }, status=500)

##################################################################################################################
def dryfilm_standard_list(request):
    # 获取所有干膜产品，按测试日期倒序排列
    dryfilmstandards = DryFilmStandard.objects.all().order_by('-product_name')
    # 记录查询结果
    logger.info(f"查询到 {len(dryfilmstandards)} 条干膜产品记录")
    return render(request, 'qc_app/dryfilm_standard/dryfilm_standard_list.html', {'dryfilmstandards': dryfilmstandards})

def create_dryfilm_standard(request):
    # 处理JSON配置表单
    config_form = DryFilmConfigForm(request.POST or None)
    main_form = DryFilmMainForm(request.POST or None, config_json=None)
    if request.method == 'POST':
        # 处理AJAX请求（JSON配置预览）
        if 'generate-json' in request.POST:

            if config_form.is_valid():
                config_data = config_form.cleaned_data
                return JsonResponse({
                    'json_data': json.dumps(config_data, cls=DjangoJSONEncoder, indent=2),
                    'success': True
                })
            return JsonResponse({'success': False})
        
        # 处理主表单提交（保存到数据库）
        if main_form.is_valid():
            # 从隐藏字段获取JSON配置
            config_json = json.loads(request.POST.get('config_json', '{}'))
            
            # 创建模型实例
            instance = main_form.save(commit=False)
            instance.config_json = config_json
            instance.save()
            
            return redirect('dryfilm_success', pk=instance.id)
    
    # 处理初始请求
    return render(request, 'qc_app/dryfilm_standard/create.html', {
        'config_form': config_form,
        'main_form': main_form
    })

def success_view(request, pk):
    from .models import DryFilmStandard
    record = DryFilmStandard.objects.get(pk=pk)
    return render(request, 'qc_app/dryfilm_standard/success.html', {'record': record})

def dryfilm_standard_edit(request, product_name):

    dryfilm_standard = get_object_or_404(DryFilmStandard, product_name=product_name)

    config_form = DryFilmConfigForm(request.POST or None)
    main_form = DryFilmMainForm(request.POST or None, config_json=None)

    if request.method == 'POST':

        # 处理AJAX请求（JSON配置预览）
        if 'generate-json' in request.POST:
            if config_form.is_valid():
                
                config_data = config_form.cleaned_data
                return JsonResponse({
                    'json_data': json.dumps(config_data, cls=DjangoJSONEncoder, indent=2),
                    'success': True
                })
            return JsonResponse({'success': False})
        
        # 处理主表单提交（保存到数据库）

        config_json = json.loads(request.POST.get('config_json', '{}'))
        
        # 假设唯一字段是'name'
        unique_field = 'product_name'
        unique_value = dryfilm_standard.product_name
        # 创建更新字典（排除唯一字段）
        update_fields = {}
        update_fields['config_json'] = config_json
        update_fields['producer'] = dryfilm_standard.producer
        
        # 使用update_or_create
        instance, created = DryFilmStandard.objects.update_or_create(
            **{unique_field: unique_value},
            defaults=update_fields
        )
            
        return redirect('dryfilm_success', pk=instance.id)
    else:
        main_form = DryFilmMainForm(instance=dryfilm_standard)

        config_form = DryFilmConfigForm(dryfilm_standard.config_json)

    return render(request, 'qc_app/dryfilm_standard/create.html', {'config_form': config_form,'main_form': main_form, 'mode': 'edit', 'dryfilm_standard': dryfilm_standard})

@require_POST
def dryfilm_standard_delete(request):
    record_id = request.POST.get('record_id')
    
    if not record_id:
        return JsonResponse({'success': False, 'message': '无效的请求参数'}, status=400)

    try:
        # 获取并删除记录
        record = get_object_or_404(DryFilmStandard, pk=record_id)
        record.delete()
        
        messages.success(request, f"{record.product_name}-{record.created_at} 已成功删除")
        # 成功响应
        return redirect('dryfilm_standard_list')
        return JsonResponse({
            'success': True, 
            'record_id': record_id,
            'message': f'批号 {record.batch_number} 已删除'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'删除失败: {str(e)}'
        }, status=500)

##################################################################################################################
# 报告中心首页
def reports_home(request):
    return render(request, 'qc_app/reports_home.html')

# Excel报告生成
def generate_excel_report(request, product_type, batch_number):
    # 获取产品数据
    product = get_object_or_404(DryFilmProduct, batch_number=batch_number)
    
    # 创建Excel工作簿
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{batch_number}_report.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "质检报告"
    
    # 添加报告头
    headers = ["质检报告", f"产品牌号: {product.product_id}", f"批号: {batch_number}"]
    for i, header in enumerate(headers, 1):
        ws.cell(row=i, column=1, value=header)
    
    # 添加数据表格
    data_rows = [
        ["检测项目", "标准要求", "实测值", "判定"],
        ["外观", "-", product.appearance, ""],
        # 其他数据行...
    ]
    
    for row_idx, row_data in enumerate(data_rows, 4):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(response)
    return response
###########################################################################################################################

def adhesive_list(request):
    products = AdhesiveProduct.objects.all()
    return render(request, 'qc_app/adhesive_list.html', {'products': products})

def adhesive_create(request):
    if request.method == 'POST':
        form = AdhesiveForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('adhesive_list')
    else:
        form = AdhesiveForm()
    
    context = {
        'form': form,
        'mode': 'create',
        'title': '添加胶粘剂产品'
    }
    return render(request, 'qc_app/adhesive_form.html', context)

def adhesive_edit(request, pk):
    product = get_object_or_404(AdhesiveProduct, pk=pk)
    
    if request.method == 'POST':
        form = AdhesiveForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            return redirect('adhesive_list')
    else:
        form = AdhesiveForm(instance=product)
    
    context = {
        'form': form,
        'mode': 'edit',
        'product': product,
        'title': '编辑胶粘剂产品'
    }
    return render(request, 'qc_app/adhesive_form.html', context)
########################################################################################################
def rawmaterial_list(request):
    materials = RawMaterial.objects.all().order_by('-received_date')
    return render(request, 'qc_app/rawmaterial_list.html', {'materials': materials})

def rawmaterial_create(request):
    if request.method == 'POST':
        form = RawMaterialForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('rawmaterial_list')
    else:
        form = RawMaterialForm()
    
    context = {
        'form': form,
        'mode': 'create',
        'title': '添加原料数据'
    }
    return render(request, 'qc_app/rawmaterial_form.html', context)

def rawmaterial_edit(request, pk):
    material = get_object_or_404(RawMaterial, pk=pk)
    
    if request.method == 'POST':
        form = RawMaterialForm(request.POST, instance=material)
        if form.is_valid():
            form.save()
            return redirect('rawmaterial_list')
    else:
        form = RawMaterialForm(instance=material)
    
    context = {
        'form': form,
        'mode': 'edit',
        'material': material,
        'title': '编辑原料数据'
    }
    return render(request, 'qc_app/rawmaterial_form.html', context)

########################################################################################################


########################################################################################################

def trialproduct_list(request):
    products = TrialProduct.objects.all().order_by('-test_date')
    return render(request, 'qc_app/trialproduct_list.html', {'products': products})

def trialproduct_create(request):
    if request.method == 'POST':
        
        form = TrialProductForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('trialproduct_list')
    else:
        form = TrialProductForm()
    
    context = {
        'form': form,
        'mode': 'create',
        'title': '添加小试产品'
    }
    return render(request, 'qc_app/trialproduct_form.html', context)

def trialproduct_edit(request, pk):
    product = get_object_or_404(TrialProduct, pk=pk)
    
    if request.method == 'POST':
        form = TrialProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            return redirect('trialproduct_list')
    else:
        form = TrialProductForm(instance=product)
    
    context = {
        'form': form,
        'mode': 'edit',
        'product': product,
        'title': '编辑小试产品'
    }
    return render(request, 'qc_app/trialproduct_form.html', context)











########################################################################################################
def report_list(request):
    reports = ReportCenter.objects.filter(
        Q(generated_by=request.user.username) | 
        Q(status__in=['approved', 'released'])
    ).order_by('-generated_date')
    
    return render(request, 'qc_app/report_list.html', {'reports': reports})










########################################################################################################
def analysis_list(request):
    analyses = AnalysisCenter.objects.all().order_by('-performed_date')
    return render(request, 'qc_app/analysis_list.html', {'analyses': analyses})












########################################################################################################
def test_permission(request):
    """测试权限页面"""
    return render(request, 'qc_app/test_permission.html', {
        'user_role': request.user_role,
        'client_ip': request.META.get('REMOTE_ADDR')
    })

# 在视图函数中添加调试输出
def test_view(request):
    print(f"用户角色: {request.user_role}")
    print(f"客户端IP: {request.META.get('REMOTE_ADDR')}")
    # ...